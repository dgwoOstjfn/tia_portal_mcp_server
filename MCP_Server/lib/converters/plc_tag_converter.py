"""
PLC Tag Table XML to Excel converter and vice versa
Handles conversion between TIA Portal PLC tag XML format and Excel format
"""
import xml.etree.ElementTree as ET
import os
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.warning("openpyxl not installed. Installing...")
    import subprocess
    import sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class PLCTagConverter:
    """Converts PLC tag tables between XML and Excel formats"""
    
    def __init__(self):
        self.namespaces = {
            'ns': 'http://www.siemens.com/automation/Openness/SW/Tags/v5'
        }
        
    def xml_to_excel(self, xml_file_path: str, output_path: str = None) -> str:
        """
        Convert PLC tag table XML to Excel format
        
        Args:
            xml_file_path: Path to input XML file
            output_path: Path for output Excel file (optional)
            
        Returns:
            Path to generated Excel file
        """
        try:
            # Parse XML
            tree = ET.parse(xml_file_path)
            root = tree.getroot()
            
            # Get table name
            tag_table = root.find('.//SW.Tags.PlcTagTable')
            if tag_table is None:
                raise ValueError("No PlcTagTable found in XML")
                
            table_name = tag_table.find('.//Name')
            table_name_str = table_name.text if table_name is not None else "UnknownTable"
            
            # Determine output path
            if output_path is None:
                base_name = Path(xml_file_path).stem
                output_dir = os.path.dirname(xml_file_path)
                output_path = os.path.join(output_dir, f"{base_name}.xlsx")
            
            # Extract different tag types
            variables = []
            user_constants = []
            system_constants = []
            
            # Process tags
            for tag in tag_table.findall('.//SW.Tags.PlcTag'):
                tag_data = self._extract_tag_data(tag)
                variables.append(tag_data)
            
            # Process user constants
            for const in tag_table.findall('.//SW.Tags.PlcUserConstant'):
                const_data = self._extract_constant_data(const)
                user_constants.append(const_data)
            
            # Process system constants (if any)
            for sys_const in tag_table.findall('.//SW.Tags.PlcSystemConstant'):
                sys_const_data = self._extract_constant_data(sys_const)
                system_constants.append(sys_const_data)
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Add sheets for each table type
            if variables:
                self._create_variables_sheet(wb, variables, table_name_str)
            
            if user_constants:
                self._create_constants_sheet(wb, user_constants, table_name_str, "User Constants")
            
            if system_constants:
                self._create_constants_sheet(wb, system_constants, table_name_str, "System Constants")
            
            # Save Excel file
            wb.save(output_path)
            
            logger.info(f"Successfully converted {xml_file_path} to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error converting XML to Excel: {e}")
            raise
    
    def _extract_tag_data(self, tag_element) -> Dict[str, Any]:
        """Extract data from a PLC tag element"""
        tag_data = {
            'Name': '',
            'DataType': '',
            'LogicalAddress': '',
            'Comment_de-DE': '',
            'Comment_en-US': '',
            'Comment_zh-CN': '',
            'Comment_hu-HU': '',
            'ExternalAccessible': 'false',
            'ExternalVisible': 'false',
            'ExternalWritable': 'false',
            'SetPoint': 'false',
            'Retain': 'false'
        }
        
        # Extract attributes
        attrs = tag_element.find('AttributeList')
        if attrs is not None:
            for attr in attrs:
                if attr.tag == 'Name':
                    tag_data['Name'] = attr.text or ''
                elif attr.tag == 'DataTypeName':
                    tag_data['DataType'] = attr.text or ''
                elif attr.tag == 'LogicalAddress':
                    tag_data['LogicalAddress'] = attr.text or ''
                elif attr.tag == 'ExternalAccessible':
                    tag_data['ExternalAccessible'] = attr.text or 'false'
                elif attr.tag == 'ExternalVisible':
                    tag_data['ExternalVisible'] = attr.text or 'false'
                elif attr.tag == 'ExternalWritable':
                    tag_data['ExternalWritable'] = attr.text or 'false'
                elif attr.tag == 'SetPoint':
                    tag_data['SetPoint'] = attr.text or 'false'
                elif attr.tag == 'Retain':
                    tag_data['Retain'] = attr.text or 'false'
        
        # Extract multilingual comments
        comment = tag_element.find('.//MultilingualText[@CompositionName="Comment"]')
        if comment is not None:
            for item in comment.findall('.//MultilingualTextItem'):
                culture = item.find('AttributeList/Culture')
                text = item.find('AttributeList/Text')
                if culture is not None and text is not None:
                    culture_str = culture.text
                    comment_text = text.text or ''
                    tag_data[f'Comment_{culture_str}'] = comment_text
        
        return tag_data
    
    def _extract_constant_data(self, const_element) -> Dict[str, Any]:
        """Extract data from a constant element"""
        const_data = {
            'Name': '',
            'DataType': '',
            'Value': '',
            'Comment_de-DE': '',
            'Comment_en-US': '',
            'Comment_zh-CN': '',
            'Comment_hu-HU': ''
        }
        
        # Extract attributes
        attrs = const_element.find('AttributeList')
        if attrs is not None:
            for attr in attrs:
                if attr.tag == 'Name':
                    const_data['Name'] = attr.text or ''
                elif attr.tag == 'DataTypeName':
                    const_data['DataType'] = attr.text or ''
                elif attr.tag == 'Value':
                    const_data['Value'] = attr.text or ''
        
        # Extract multilingual comments
        comment = const_element.find('.//MultilingualText[@CompositionName="Comment"]')
        if comment is not None:
            for item in comment.findall('.//MultilingualTextItem'):
                culture = item.find('AttributeList/Culture')
                text = item.find('AttributeList/Text')
                if culture is not None and text is not None:
                    culture_str = culture.text
                    comment_text = text.text or ''
                    const_data[f'Comment_{culture_str}'] = comment_text
        
        return const_data
    
    def _create_variables_sheet(self, wb: Workbook, variables: List[Dict], table_name: str):
        """Create Excel sheet for variables"""
        if not variables:
            return
            
        # Create sheet
        ws = wb.create_sheet(title="Variables")
        
        # Define column headers
        headers = ['Name', 'DataType', 'LogicalAddress', 
                  'Comment_de-DE', 'Comment_en-US', 'Comment_zh-CN', 'Comment_hu-HU',
                  'ExternalAccessible', 'ExternalVisible', 'ExternalWritable', 
                  'SetPoint', 'Retain']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add table info
        ws['A1'] = f"PLC Tag Table: {table_name}"
        ws['A2'] = "Type: Variables"
        ws.merge_cells('A1:L1')
        ws.merge_cells('A2:L2')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].font = Font(bold=True, size=12)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row_idx, tag_data in enumerate(variables, 5):
            for col_idx, header in enumerate(headers, 1):
                value = tag_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_constants_sheet(self, wb: Workbook, constants: List[Dict], table_name: str, sheet_title: str):
        """Create Excel sheet for constants"""
        if not constants:
            return
            
        # Create sheet
        ws = wb.create_sheet(title=sheet_title)
        
        # Define column headers
        headers = ['Name', 'DataType', 'Value',
                  'Comment_de-DE', 'Comment_en-US', 'Comment_zh-CN', 'Comment_hu-HU']
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add table info
        ws['A1'] = f"PLC Tag Table: {table_name}"
        ws['A2'] = f"Type: {sheet_title}"
        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws['A1'].font = Font(bold=True, size=14)
        ws['A2'].font = Font(bold=True, size=12)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row_idx, const_data in enumerate(constants, 5):
            for col_idx, header in enumerate(headers, 1):
                value = const_data.get(header, '')
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def excel_to_xml(self, excel_file_path: str, output_path: str = None, 
                     table_name: str = None, engineering_version: str = "V17") -> str:
        """
        Convert Excel file back to PLC tag table XML format
        
        Args:
            excel_file_path: Path to Excel file to convert
            output_path: Path for output XML file (optional)
            table_name: Name for the tag table (optional)
            engineering_version: TIA Portal version (default: V17)
            
        Returns:
            Path to generated XML file
        """
        try:
            # Load Excel workbook
            wb = openpyxl.load_workbook(excel_file_path, data_only=True)
            
            # Determine table name
            if table_name is None:
                # Try to extract from first sheet's info or use file name
                base_name = Path(excel_file_path).stem
                table_name = base_name
            
            # Create XML structure
            root = ET.Element("Document")
            
            # Add engineering version
            eng = ET.SubElement(root, "Engineering")
            eng.set("version", engineering_version)
            
            # Add DocumentInfo
            doc_info = ET.SubElement(root, "DocumentInfo")
            created = ET.SubElement(doc_info, "Created")
            from datetime import datetime
            created.text = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%S.%fZ")
            export_setting = ET.SubElement(doc_info, "ExportSetting")
            export_setting.text = "None"
            
            # Add SW.Tags.PlcTagTable
            tag_table = ET.SubElement(root, "SW.Tags.PlcTagTable")
            tag_table.set("ID", "0")
            
            attr_list = ET.SubElement(tag_table, "AttributeList")
            name_elem = ET.SubElement(attr_list, "Name")
            name_elem.text = table_name
            
            object_list = ET.SubElement(tag_table, "ObjectList")
            
            # Process each sheet in Excel file
            element_id = 1
            for sheet_name in wb.sheetnames:
                element_id = self._process_excel_sheet(wb[sheet_name], object_list, element_id)
            
            # Generate output path if not provided
            if output_path is None:
                output_dir = os.path.dirname(excel_file_path)
                output_path = os.path.join(output_dir, f"{table_name}.xml")
            
            # Write XML with proper formatting
            self._write_formatted_xml(root, output_path)
            
            logger.info(f"Successfully converted Excel file to {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error converting Excel to XML: {e}")
            raise
    
    def _process_excel_sheet(self, ws, object_list, element_id: int) -> int:
        """Process a single Excel sheet and add elements to object list"""
        # Skip if sheet is empty
        if ws.max_row < 4:
            return element_id
        
        # Get headers from row 4
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=4, column=col).value
            if header:
                headers.append(header)
        
        if not headers:
            return element_id
        
        # Process data rows starting from row 5
        for row in range(5, ws.max_row + 1):
            row_data = {}
            for col, header in enumerate(headers, 1):
                value = ws.cell(row=row, column=col).value
                if value is not None:
                    row_data[header] = str(value)
            
            # Skip empty rows
            if not any(row_data.values()):
                continue
            
            # Check if it's variables or constants
            if 'LogicalAddress' in headers:
                # Variables
                element_id = self._add_tag_element(object_list, row_data, element_id)
            elif 'Value' in headers:
                # Constants
                element_id = self._add_constant_element(object_list, row_data, element_id)
        
        return element_id
    
    def _add_tag_element(self, object_list, row: Dict, element_id: int) -> int:
        """Add a PLC tag element to the object list"""
        tag = ET.SubElement(object_list, "SW.Tags.PlcTag")
        tag.set("ID", hex(element_id)[2:].upper())
        tag.set("CompositionName", "Tags")
        
        attr_list = ET.SubElement(tag, "AttributeList")
        
        # Add attributes
        if row.get('DataType'):
            elem = ET.SubElement(attr_list, "DataTypeName")
            elem.text = row['DataType']
        
        if row.get('ExternalAccessible'):
            elem = ET.SubElement(attr_list, "ExternalAccessible")
            elem.text = row['ExternalAccessible']
        
        if row.get('LogicalAddress'):
            elem = ET.SubElement(attr_list, "LogicalAddress")
            elem.text = row['LogicalAddress']
        
        if row.get('Name'):
            elem = ET.SubElement(attr_list, "Name")
            elem.text = row['Name']
        
        # Add optional attributes if present and not default
        if row.get('ExternalVisible') and row['ExternalVisible'] != 'false':
            elem = ET.SubElement(attr_list, "ExternalVisible")
            elem.text = row['ExternalVisible']
        
        if row.get('ExternalWritable') and row['ExternalWritable'] != 'false':
            elem = ET.SubElement(attr_list, "ExternalWritable")
            elem.text = row['ExternalWritable']
        
        if row.get('SetPoint') and row['SetPoint'] != 'false':
            elem = ET.SubElement(attr_list, "SetPoint")
            elem.text = row['SetPoint']
        
        if row.get('Retain') and row['Retain'] != 'false':
            elem = ET.SubElement(attr_list, "Retain")
            elem.text = row['Retain']
        
        element_id += 1
        
        # Add comments if present
        comments = {}
        for key in row:
            if key.startswith('Comment_'):
                culture = key.replace('Comment_', '')
                if row[key]:  # Only add non-empty comments
                    comments[culture] = row[key]
        
        if comments:
            obj_list = ET.SubElement(tag, "ObjectList")
            comment_elem = ET.SubElement(obj_list, "MultilingualText")
            comment_elem.set("ID", hex(element_id)[2:].upper())
            comment_elem.set("CompositionName", "Comment")
            element_id += 1
            
            comment_obj_list = ET.SubElement(comment_elem, "ObjectList")
            
            for culture, text in comments.items():
                item = ET.SubElement(comment_obj_list, "MultilingualTextItem")
                item.set("ID", hex(element_id)[2:].upper())
                item.set("CompositionName", "Items")
                element_id += 1
                
                item_attr_list = ET.SubElement(item, "AttributeList")
                culture_elem = ET.SubElement(item_attr_list, "Culture")
                culture_elem.text = culture
                text_elem = ET.SubElement(item_attr_list, "Text")
                text_elem.text = text
        
        return element_id
    
    def _add_constant_element(self, object_list, row: Dict, element_id: int) -> int:
        """Add a constant element to the object list"""
        const = ET.SubElement(object_list, "SW.Tags.PlcUserConstant")
        const.set("ID", hex(element_id)[2:].upper())
        const.set("CompositionName", "UserConstants")
        
        attr_list = ET.SubElement(const, "AttributeList")
        
        # Add attributes
        if row.get('DataType'):
            elem = ET.SubElement(attr_list, "DataTypeName")
            elem.text = row['DataType']
        
        if row.get('Name'):
            elem = ET.SubElement(attr_list, "Name")
            elem.text = row['Name']
        
        if row.get('Value'):
            elem = ET.SubElement(attr_list, "Value")
            elem.text = row['Value']
        
        element_id += 1
        
        # Add comments if present
        comments = {}
        for key in row:
            if key.startswith('Comment_'):
                culture = key.replace('Comment_', '')
                if row[key]:  # Only add non-empty comments
                    comments[culture] = row[key]
        
        if comments:
            obj_list = ET.SubElement(const, "ObjectList")
            comment_elem = ET.SubElement(obj_list, "MultilingualText")
            comment_elem.set("ID", hex(element_id)[2:].upper())
            comment_elem.set("CompositionName", "Comment")
            element_id += 1
            
            comment_obj_list = ET.SubElement(comment_elem, "ObjectList")
            
            for culture, text in comments.items():
                item = ET.SubElement(comment_obj_list, "MultilingualTextItem")
                item.set("ID", hex(element_id)[2:].upper())
                item.set("CompositionName", "Items")
                element_id += 1
                
                item_attr_list = ET.SubElement(item, "AttributeList")
                culture_elem = ET.SubElement(item_attr_list, "Culture")
                culture_elem.text = culture
                text_elem = ET.SubElement(item_attr_list, "Text")
                text_elem.text = text
        
        return element_id
    
    def _write_formatted_xml(self, root, output_path: str):
        """Write XML with proper formatting"""
        # Convert to string and format
        from xml.dom import minidom
        
        rough_string = ET.tostring(root, encoding='unicode')
        reparsed = minidom.parseString(rough_string)
        
        # Get pretty printed version
        pretty_xml = reparsed.toprettyxml(indent="  ", encoding='utf-8')
        
        # Remove extra blank lines
        lines = pretty_xml.decode('utf-8').split('\n')
        non_empty_lines = [line for line in lines if line.strip()]
        
        # Write to file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(non_empty_lines))